#import openpyxl and speech_recognization module
import openpyxl
import speech_recognition as sr

#works as recogizer to recognize our audio
r = sr.Recognizer()

#To open the workbook,workbook object is created
wb = openpyxl.load_workbook("Top10ranksheet.xlsx")
ws=wb.active

#initializing the source to the microphone
with sr.Microphone() as source:
    r.pause_threshold = 1
    r.adjust_for_ambient_noise(source)
    print("Speak Anything :")
    audio = r.listen(source)
    try:
        text = r.recognize_google(audio)
        print("You said : {}".format(text))
    except:
        print("Sorry could not recognize what you said")


search=r.recognize_google(audio)

for r in range(1,ws.max_row+1):
    for c in range(1,ws.max_column+1):
       
        s = str(ws.cell(r,c).value)
        cell_up=ws.cell(row=r,column=6)
        
        if  str(s)==search: 
            cell_up.value="yes"
            print("updation done")
        
            
wb.save('Top10ranksheet.xlsx')


