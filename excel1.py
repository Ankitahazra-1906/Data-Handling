# Python program to read an excel file

# import openpyxl module
import openpyxl

# Give the location of the file
#path = "C:\Users\User\Documents\Top10ranksheet.xlsx"

# To open the workbook
# workbook object is created
wb_obj = openpyxl.load_workbook('Top10ranksheet.xlsx')

# Get workbook active sheet object
# from the active attribute
sheet_obj = wb_obj.active

# Cell objects also have a row, column,
# and coordinate attributes that provide
# location information for the cell.

# Note: The first row or
# column integer is 1, not 0.

# Cell object is created by using
# sheet object's cell() method.
x=int(input("Enter the row no: "))
y=int(input("Enter the column no: "))
cell_obj = sheet_obj.cell(row=x,  column=y)

# Print value of cell object
# using the value attribute
print("The required cell is:",cell_obj.value)


#print the total no of rows
print("The total no of rows are: ",sheet_obj.max_row)

#print the total no of coloumns
print("The total no of columns are: ",sheet_obj.max_column)

max_col = sheet_obj.max_column
max_row = sheet_obj.max_row
 
# Loop will print all columns name
print("All column names are listed below: ")
for i in range(1, max_col + 1):
    cell_obj = sheet_obj.cell(row = 1, column = i)
    print(cell_obj.value)

# Loop will print the first column name
z=int(input("Enter the column no you want to print: "))
print("Column no",z, "is listed below: ")
for i in range(1, max_row + 1):
    cell_obj = sheet_obj.cell(row = i, column = z)
    print(cell_obj.value)


# Will print a particular row value
p=int(input("Enter the row no you want to print: "))
print("Row no",p, "is listed below: ")
for i in range(1, max_col + 1):
    cell_obj = sheet_obj.cell(row = p, column = i)
    print(cell_obj.value, end = " ")